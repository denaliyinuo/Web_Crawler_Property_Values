from selenium import webdriver
from openpyxl import load_workbook
import bs4
import requests as re

# path = '/Users/nathanoliver/Desktop/Python/Anchorage Property Information/anchorage_property_information.xlsx'
path = '/Users/nathanoliver/Desktop/Property Web Scraper/MoA_Property_Information_empty.xlsx'

print('loading workbook')
wb = load_workbook(path)
print('loaded workbook')

sheet = wb["Sheet1"]

sheet.cell(row=1, column=1).value = 'Lot Number'
sheet.cell(row=1, column=2).value = 'Lot Type'
sheet.cell(row=1, column=3).value = 'Address'
sheet.cell(row=1, column=4).value = 'Lot Size'
sheet.cell(row=1, column=5).value = 'Zone'
sheet.cell(row=1, column=6).value = '2018 Land Price'
sheet.cell(row=1, column=7).value = '2018 Building Price'
sheet.cell(row=1, column=8).value = '2018 Total Price'
sheet.cell(row=1, column=9).value = '2019 Land Price'
sheet.cell(row=1, column=10).value = '2019 Building Price'
sheet.cell(row=1, column=11).value = '2019 Total Price'
sheet.cell(row=1, column=12).value = '2020 Land Price'
sheet.cell(row=1, column=13).value = '2020 Building Price'
sheet.cell(row=1, column=14).value = '2020 Total Price'
sheet.cell(row=1, column=15).value = 'Living Area'
sheet.cell(row=1, column=16).value = 'Style'
sheet.cell(row=1, column=17).value = 'Story Height'
sheet.cell(row=1, column=18).value = 'Tax District'
sheet.cell(row=1, column=19).value = 'Year Built'
sheet.cell(row=1, column=20).value = 'Remodeled'
sheet.cell(row=1, column=21).value = 'Grid'
sheet.cell(row=1, column=22).value = 'Total Rooms'
sheet.cell(row=1, column=23).value = 'Bedrooms'
sheet.cell(row=1, column=24).value = 'Full Baths'
sheet.cell(row=1, column=25).value = 'Half Baths'
sheet.cell(row=1, column=26).value = 'Heat Type'
sheet.cell(row=1, column=27).value = 'Fuel Type'
sheet.cell(row=1, column=28).value = 'FP Stacks'
sheet.cell(row=1, column=29).value = 'Grade'
sheet.cell(row=1, column=30).value = 'Condition'
sheet.cell(row=1, column=31).value = 'Res/Com Type'


def text_extraction(html):
    info = browser.find_elements_by_xpath(html)
    for i in info:
        return i.text


def text_creator(info):
    for i in info:
        return i.text


def split_text(info):
    test = info.split()
    return test


def split_text_by(info):
    data = info.split('-')
    return data


browser = webdriver.Chrome()
website = 'https://www.muni.org/pw/public.html'
browser.get(website)

split_next_house = ['010', '352', '14', '000']


row_number = 1
house_number = -1
n = 0

while split_next_house[0] != '100':
    house_number += 1

    input1 = browser.find_element_by_name('PAR1')
    input2 = browser.find_element_by_name('PAR2')
    input3 = browser.find_element_by_name('PAR3')
    input4 = browser.find_element_by_name('APP')

    input1.send_keys(split_next_house[0])
    input2.send_keys(split_next_house[1])
    input3.send_keys(split_next_house[2])
    input4.send_keys(split_next_house[3])

    html = '/html/body/form[2]/table/tbody/tr[2]/td/table[2]/tbody/tr[2]/td[2]/img'
    submit_button = browser.find_elements_by_xpath(html)[0]
    submit_button.click()

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[2]/td[3]'
    address = browser.find_elements_by_xpath(html)

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[3]/td[1]/a'
    next_house = browser.find_elements_by_xpath(html)

    address = text_creator(address)
    next_house = text_creator(next_house)
    print(next_house)
    split_next_house = split_text_by(next_house)
    print(split_next_house)

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[2]/td[1]/a'
    submit_button = browser.find_elements_by_xpath(html)[0]
    submit_button.click()

    xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[2]/td/pre/span'
    residential_headline_text = text_extraction(xpath)
    print(residential_headline_text)
    res_data = split_text(residential_headline_text)
    print(res_data)

    if len(res_data) > 2:

        row_number += 1

        print('entered Residential')

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[2]/td/pre/span[2]'
        parcel_number = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[2]'
        lot_size = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[3]/span[1]'
        zone = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[6]/td/pre[1]/span/span'

        prices = text_extraction(xpath)
        prices = split_text(prices)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[10]/td/pre/span[7]/span[6]/span[7]'
        living_area = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[2]'
        style = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[4]'
        story_height = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[3]/span[11]/span[2]'
        tax_district = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[2]'
        year_built = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[4]'
        remodeled = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[3]/span[11]/span[14]/span[2]'
        grid = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[2]'
        total_rooms = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[4]'
        bedrooms = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[2]'
        full_baths = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[4]'
        half_baths = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[6]/span[2]'
        heat_type = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[6]/span[4]'
        fuel_type = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[6]/span[6]/span[2]'
        fp_stacks = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[2]'
        grade = text_extraction(xpath)

        xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[8]/td/pre/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]/span[6]'
        condition = text_extraction(xpath)

        sheet.cell(row=row_number, column=1).value = res_data[1]
        sheet.cell(row=row_number, column=2).value = res_data[2]
        sheet.cell(row=row_number, column=3).value = address
        sheet.cell(row=row_number, column=4).value = lot_size
        sheet.cell(row=row_number, column=5).value = zone
        sheet.cell(row=row_number, column=6).value = prices[0]
        sheet.cell(row=row_number, column=7).value = prices[1]
        sheet.cell(row=row_number, column=8).value = prices[2]
        sheet.cell(row=row_number, column=9).value = prices[6]
        sheet.cell(row=row_number, column=10).value = prices[7]
        sheet.cell(row=row_number, column=11).value = prices[8]
        sheet.cell(row=row_number, column=12).value = prices[13]
        sheet.cell(row=row_number, column=13).value = prices[14]
        sheet.cell(row=row_number, column=14).value = prices[15]
        sheet.cell(row=row_number, column=15).value = living_area
        sheet.cell(row=row_number, column=16).value = style
        sheet.cell(row=row_number, column=17).value = story_height
        sheet.cell(row=row_number, column=18).value = tax_district
        sheet.cell(row=row_number, column=19).value = year_built
        sheet.cell(row=row_number, column=20).value = remodeled
        sheet.cell(row=row_number, column=21).value = grid
        sheet.cell(row=row_number, column=22).value = total_rooms
        sheet.cell(row=row_number, column=23).value = bedrooms
        sheet.cell(row=row_number, column=24).value = full_baths
        sheet.cell(row=row_number, column=25).value = half_baths
        sheet.cell(row=row_number, column=26).value = heat_type
        sheet.cell(row=row_number, column=27).value = fuel_type
        sheet.cell(row=row_number, column=28).value = fp_stacks
        sheet.cell(row=row_number, column=29).value = grade
        sheet.cell(row=row_number, column=30).value = condition
        sheet.cell(row=row_number, column=31).value = res_data[3]

        n += 1

        if n == 1:
            print('')
            print('Saving...')

            wb.save(path)

            print('')
            print('*****Export Complete!*****')

            n = 0

    else:
        pass

    html = '//*[@id="MOA_BreadCrumbControl"]/a[4]'
    return_button = browser.find_elements_by_xpath(html)[0]
    return_button.click()
    enter = 0

browser.close()
