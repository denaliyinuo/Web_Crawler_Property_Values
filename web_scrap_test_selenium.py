from selenium import webdriver
from openpyxl import load_workbook
import bs4
import requests as re

# path = '/Users/nathanoliver/Desktop/Python/Anchorage Property Information/anchorage_property_information.xlsx'
path = '/Users/nathanoliver/Desktop/Python/Anchorage Property Information/test.xlsx'

print('loading workbook')
wb = load_workbook(path)
print('loaded workbook')

sheet = wb["Sheet1"]

sheet.cell(row=1, column=1).value = 'Lot Number'
sheet.cell(row=1, column=2).value = 'Lot Type'
sheet.cell(row=1, column=3).value = 'Address'
sheet.cell(row=1, column=4).value = 'Lot Size'
sheet.cell(row=1, column=5).value = 'Zone'
sheet.cell(row=1, column=6).value = '2018 Land Price'
sheet.cell(row=1, column=7).value = '2018 Building Price'
sheet.cell(row=1, column=8).value = '2018 Total Price'
sheet.cell(row=1, column=9).value = '2019 Land Price'
sheet.cell(row=1, column=10).value = '2019 Building Price'
sheet.cell(row=1, column=11).value = '2019 Total Price'
sheet.cell(row=1, column=12).value = '2020 Land Price'
sheet.cell(row=1, column=13).value = '2020 Building Price'
sheet.cell(row=1, column=14).value = '2020 Total Price'
sheet.cell(row=1, column=15).value = 'Living Area'


def text_extraction(html):
    info = browser.find_elements_by_xpath(html)
    for i in info:
        return i.text


def text_creator(info):
    for i in info:
        return i.text


def split_text(info, i):
    test = info[i].split()
    return test


def split_text_by(info):
    data = info.split('-')
    return data


def split_text_2(info):
    j = 0
    for i in info:

        print(j)
        print(i.text)
        info_collect.insert(j, i.text)

        j = j + 1

    return info1


browser = webdriver.Chrome()
website = 'https://www.muni.org/pw/public.html'
browser.get(website)

row_number = 1

for house_number in range(3000):

    input1 = browser.find_element_by_name('PAR1')
    input2 = browser.find_element_by_name('PAR2')
    input3 = browser.find_element_by_name('PAR3')

    if house_number == 0:
        input1.send_keys("050")
        input2.send_keys("021")
        input3.send_keys("01")

        # input1.send_keys("011")
        # input2.send_keys("154")
        # input3.send_keys("07")

    else:
        input1.send_keys(split_next_house[0])
        input2.send_keys(split_next_house[1])
        input3.send_keys(split_next_house[2])

    html = '/html/body/form[2]/table/tbody/tr[2]/td/table[2]/tbody/tr[2]/td[2]/img'
    submit_button = browser.find_elements_by_xpath(html)[0]
    submit_button.click()

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[2]/td[3]'
    address = browser.find_elements_by_xpath(html)

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[3]/td[1]/a'
    next_house = browser.find_elements_by_xpath(html)

    address = text_creator(address)
    next_house = text_creator(next_house)
    split_next_house = split_text_by(next_house)

    # sheet.cell(row=house_number + 2, column=3).value = address

    html = '/html/body/table/tbody/tr[2]/td/table[5]/tbody/tr[2]/td[1]/a'
    submit_button = browser.find_elements_by_xpath(html)[0]
    submit_button.click()

    info1 = browser.find_elements_by_class_name('T')
    info2 = browser.find_elements_by_class_name('Y')
    info3 = browser.find_elements_by_class_name('B')
    # html = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[2]'

    # lot_size = browser.find_elements_by_xpath(html)

    # for i in lot_size:
    #     print(i.text)

    xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[2]'
    lot_size = text_extraction(xpath)

    xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[4]/td/pre[2]/span[3]/span[1]'
    zone = text_extraction(xpath)

    # xpath = '/html/body/form[4]/table/tbody/tr[3]/td/table[4]/tbody/tr[6]/td/pre[1]/span/span/span[3]/span[1]'
    # path = browser.find_elements_by_xpath(xpath)
    # prices_2020 = text_creator(path)

    # print(prices_2020)
    # prices_2020_split = split_text(prices_2020, i=0)
    # print(prices_2020_split)

    # print(zone)

    j = 0

    info_collect = ['0']
    address_collect = ['0']

    for i in info1:

        # print(j)
        # print(i.text)
        info_collect.insert(j, i.text)

        j = j + 1

    # for i in info2:
    #     print('info2: ' + str(i))
    #     print(i.text)

    for i in info3:

        address_collect.insert(j, i.text)
        # print(i.text)

    for i in range(len(info_collect)):

        test = split_text(info_collect, i)
        # print(test)
        if i == 0:
            if len(test) > 2:
                if test[3] == '1-Family':
                    row_number = row_number + 1
                    print(row_number)
                    # lot number
                    # print(row_number)
                    # print(test[1])
                    # print(test[2])
                    # print(address)
                    sheet.cell(row=row_number, column=1).value = test[1]
                    sheet.cell(row=row_number, column=2).value = test[2]
                    sheet.cell(row=row_number, column=3).value = address
                    sheet.cell(row=row_number, column=4).value = lot_size
                    sheet.cell(row=row_number, column=5).value = zone
                    enter = 1
        if i == 2:
            for k in range(len(test)):
                pass
                # print(k)
                # print(test[k])
        if i == 17:
            if enter == 1:
                # for k in range(len(test)):
                    # print(test[6])
                sheet.cell(row=row_number, column=6).value = test[6]
                sheet.cell(row=row_number, column=7).value = test[7]
                sheet.cell(row=row_number, column=8).value = test[8]
                sheet.cell(row=row_number, column=9).value = test[12]
                sheet.cell(row=row_number, column=10).value = test[13]
                sheet.cell(row=row_number, column=11).value = test[14]
                sheet.cell(row=row_number, column=12).value = test[19]
                sheet.cell(row=row_number, column=13).value = test[20]
                sheet.cell(row=row_number, column=14).value = test[21]
        if i == 61:
            if enter == 1:
                print(test[-1])
                sheet.cell(row=row_number, column=15).value = test[-1]

    print('')
    print('Saving...')

    wb.save(path)

    print('')
    print('*****Export Complete!*****')
    html = '//*[@id="MOA_BreadCrumbControl"]/a[4]'
    return_button = browser.find_elements_by_xpath(html)[0]
    return_button.click()

    del(info1)
    del(info2)
    del(info3)
    del(info_collect)
    del(test)
    enter = 0

browser.close()
